import os

from django.conf import settings
from django.core import management
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from data_bridge_app.models import (
    ECV,
    Ai,
    AiCategory,
    AiType,
    AiUse,
    Dataset,
    DatasetProvider,
    Filter,
    Project,
    Relationship,
    RelationType,
)

ID_1 = 0
START_DATE_1 = 1
END_DATE_1 = 2
FILTER_1 = 3
PROVIDER_1 = 4
ECV_1 = 5
RELATIONSHIP_1 = 6
RELATIONSHIP_2 = 7
ID_2 = 8
FILTER_2 = 9
PROVIDER_2 = 10
ECV_2 = 11
DESCRIPTION = 12

AI_PROVIDER = 0
AI_PROJECT = 1
AI_DATASET = 2
AI_RELATIONSHIP_1 = 3
AI_RELATIONSHIP_2 = 4
AI_TYPE = 5
AI_CATEGORY = 6
AI_USE = 7
AI_DESCRIPTION = 8


def get_from_github(dir: str, branch: str = "main"):
    import requests

    print(f"Downloading from Github: {branch}")
    link = f"https://github.com/cedadev/cci_data_bridge_inputs/raw/refs/heads/{branch}/EEE2000-metadata_mapping.xlsx"

    resp = requests.get(link)
    with open(f"{dir}/testfile.xlsx", "wb") as f:
        f.write(resp.content)

    return f"{dir}/testfile.xlsx"


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument(
            "--excel_wb",
            dest="excel_wb",
            type=str,
            help="The excel workbook to import",
            default=None,
            nargs="?",
        )
        parser.add_argument(
            "--dir", dest="dir", type=str, help="Temp Directory", default="/tmp"
        )
        parser.add_argument(
            "--branch",
            dest="branch",
            type=str,
            help="Data Inputs branch",
            default="main",
        )

    def handle(self, dir: str, excel_wb: str | None, branch: str, **kwargs):

        if not os.path.isfile(str(settings.DATABASES["default"]["NAME"])):
            print(
                f"[info] Migrating database to {str(settings.DATABASES['default']['NAME'])}"
            )
            management.call_command("migrate", interactive=False)
        else:
            print(
                f"[info] Skipping migration for database at {str(settings.DATABASES['default']['NAME'])}"
            )

        print("[info] Import data from spreadsheet")
        if not excel_wb:
            excel_wb = get_from_github(dir, branch)
        _clean()
        w_book = load_workbook(filename=excel_wb)
        related_types = _write_related_types(w_book)
        w_sheet = w_book["Mapping"]
        providers = _write_providers()
        ecvs = _write_ecvs(w_sheet)
        filters = _write_filters(w_sheet)
        _write_datasets(w_sheet, ecvs, filters, providers, related_types)

        ai_w_sheet = w_book["AI Mapping"]
        _write_ais(ai_w_sheet)

        print("Database updated")


def _clean():
    Dataset.objects.all().delete()
    ECV.objects.all().delete()
    Filter.objects.all().delete()
    DatasetProvider.objects.all().delete()
    RelationType.objects.all().delete()
    Relationship.objects.all().delete()
    Ai.objects.all().delete()
    AiType.objects.all().delete()
    AiCategory.objects.all().delete()
    AiUse.objects.all().delete()


def _write_related_types(w_book):
    w_sheet = w_book.get_sheet_by_name("Relationship definitions")
    related_types = {}
    for row in w_sheet.iter_rows(min_row=4, max_col=4, max_row=w_sheet.max_row):
        name = row[0].value
        description = row[1].value
        if description is None:
            description = ""
        related_types[name] = RelationType.objects.create(
            name=name, description=description
        )

    return related_types


def _write_providers():
    providers = {}
    providers["C3S Climate Data Store"] = DatasetProvider.objects.create(
        name="C3S Climate Data Store"
    )
    providers["CCI Open Data Portal"] = DatasetProvider.objects.create(
        name="CCI Open Data Portal"
    )
    providers["CCI Archive on CEDA"] = DatasetProvider.objects.create(
        name="CCI Archive on CEDA"
    )
    providers["OSI SAF"] = DatasetProvider.objects.create(name="OSI SAF")
    providers["CM SAF"] = DatasetProvider.objects.create(name="CM SAF")
    providers["RECCAP-2"] = DatasetProvider.objects.create(name="RECCAP-2")
    return providers


def _write_ecvs(w_sheet):
    ecvs_from_ws = _get_ecvs(w_sheet)
    ecvs = {}
    for ecv in ecvs_from_ws:
        ecvs[ecv] = ECV.objects.create(name=ecv)

    return ecvs


def _get_ecvs(w_sheet):
    ecvs = set()
    for row in w_sheet.iter_rows(min_row=3, max_col=12, max_row=w_sheet.max_row):
        value = row[ECV_1].value
        if value is not None:
            ecvs.update(_get_values(value))
        value = row[ECV_2].value
        if value is not None:
            ecvs.update(_get_values(value))
    return ecvs


def _write_filters(w_sheet):
    filters_from_ws = _get_filters(w_sheet)
    filters = {}
    for filter_ in filters_from_ws:
        bits = filter_.split("=")
        filters[filter_] = Filter.objects.create(name=bits[0], value=bits[1])

    return filters


def _get_filters(w_sheet):
    filters = set()
    for row in w_sheet.iter_rows(min_row=3, max_col=12, max_row=w_sheet.max_row):
        value = row[FILTER_1].value
        if value is not None:
            filters.update(_get_values(value))
        value = row[FILTER_2].value
        if value is not None:
            filters.update(_get_values(value, "drs="))
    return filters


def _write_ais(w_sheet):
    for row in w_sheet.iter_rows(min_row=3, max_col=13, max_row=w_sheet.max_row):
        ai_category, _ = AiCategory.objects.get_or_create(name=row[AI_CATEGORY].value)
        ai_type, _ = AiType.objects.get_or_create(
            name=row[AI_TYPE].value, category=ai_category
        )
        ai_use, _ = AiUse.objects.get_or_create(name=row[AI_USE].value)
        ai, _ = Ai.objects.get_or_create(use=ai_use, type=ai_type)

        provider, _ = DatasetProvider.objects.get_or_create(name=row[AI_PROVIDER].value)
        relationship_d_1 = None
        relationship_d_2 = None
        if row[AI_DATASET].value.strip() != "-":
            dataset, _ = Dataset.objects.get_or_create(
                url=row[AI_DATASET].value, dataset_provider=provider
            )

            relationship_d_1 = Relationship.objects.create(
                source=dataset,
                target=ai,
                description=row[AI_DESCRIPTION].value or "",
            )

            relationship_d_2 = Relationship.objects.create(
                source=ai,
                target=dataset,
                description=row[AI_DESCRIPTION].value or "",
            )

        project, _ = Project.objects.get_or_create(
            name=row[AI_PROJECT].value,
            dataset_provider=provider,
        )

        relationship_p_1 = Relationship.objects.create(
            source=project,
            target=ai,
            description=row[AI_DESCRIPTION].value or "",
        )

        relationship_p_2 = Relationship.objects.create(
            source=ai,
            target=project,
            description=row[AI_DESCRIPTION].value or "",
        )

        for rel_type in _get_values(row[AI_RELATIONSHIP_1].value):
            relationship_type, _ = RelationType.objects.get_or_create(name=rel_type)
            relationship_p_1.relationships.add(relationship_type)
            if relationship_d_1:
                relationship_d_1.relationships.add(relationship_type)

        for rel_type in _get_values(row[AI_RELATIONSHIP_2].value):
            relationship_type, _ = RelationType.objects.get_or_create(name=rel_type)
            relationship_p_2.relationships.add(relationship_type)
            if relationship_d_2:
                relationship_d_2.relationships.add(relationship_type)


def _write_datasets(w_sheet, ecvs, filters, providers, related_types):
    for row in w_sheet.iter_rows(min_row=3, max_col=13, max_row=w_sheet.max_row):
        if row[ID_1].value is None:
            continue

        # create primary dataset
        ds_1 = Dataset.objects.create(
            url=row[ID_1].value,
            dataset_provider=providers[row[PROVIDER_1].value],
        )

        # add start date
        if (
            row[START_DATE_1].value is not None
            and str(row[START_DATE_1].value).strip() != "-"
        ):
            ds_1.start_date = str(row[START_DATE_1].value).split(" ")[0]
            ds_1.start_date = row[START_DATE_1].value
            ds_1.save()

        # add end date
        if (
            row[END_DATE_1].value is not None
            and str(row[START_DATE_1].value).strip() != "-"
        ):
            ds_1.end_date = row[END_DATE_1].value
            # print(x, ds_1.end_date)
            ds_1.save()

        # add ECVs
        if row[ECV_1].value is not None:
            for ecv in _get_values(row[ECV_1].value):
                ds_1.ecvs.add(ecvs[ecv])

        # add filters
        for filter_ in _get_values(row[FILTER_1].value):
            ds_1.filters.add(filters[filter_])

        if row[ID_2].value is None:
            continue

        # check if secondary dataset exists
        matching_datasets = _get_existing_ds(
            row[ID_2].value, _get_values(row[FILTER_2].value, "drs=")
        )
        if len(matching_datasets.all()) == 1:
            ds_2 = matching_datasets[0]
        else:
            # create secondary dataset
            ds_2 = Dataset.objects.create(
                url=row[ID_2].value,
                dataset_provider=providers[row[PROVIDER_2].value],
                start_date="2015-01-01",  # TODO
                end_date="2023-01-01",  # TODO
            )

        # add ECVs
        if row[ECV_2].value is not None:
            for ecv in _get_values(row[ECV_2].value):
                ds_2.ecvs.add(ecvs[ecv])

        # add filters
        for filter_ in _get_values(row[FILTER_2].value, "drs="):
            ds_2.filters.add(filters[f"{filter_}"])

        # add relationships
        relationship_types = []
        for relationship in _get_values(row[RELATIONSHIP_1].value):
            if "/" in relationship:
                # not sure of relationship to use, ignore for now
                print(f"Found '/' in {relationship}, ignoring")
                continue
            relationship_types.append(related_types[relationship])

        if len(relationship_types) > 0:
            description = row[DESCRIPTION].value
            if description is None:
                description = ""
            rel = Relationship.objects.create(
                source=ds_1,
                # relationships=relationship_types,
                target=ds_2,
                description=description,
            )
            for relationship_type in relationship_types:
                rel.relationships.add(relationship_type)

        if row[RELATIONSHIP_2].value is None:
            # no relationships yet
            continue

        relationship_types = []
        for relationship in _get_values(row[RELATIONSHIP_2].value):
            if relationship is None:
                continue
            if "/" in relationship:
                print(f"Found '/' in {relationship}, ignoring")
                continue
            relationship_types.append(related_types[relationship])

        if len(relationship_types) > 0:
            description = row[DESCRIPTION].value
            if description is None:
                description = ""
            rel = Relationship.objects.create(
                source=ds_2,
                target=ds_1,
                description=description,
            )
            for relationship_type in relationship_types:
                rel.relationships.add(relationship_type)


def _get_existing_ds(url, filters):
    results = Dataset.objects.all()

    # get the datasets for the given url
    results = results.filter(url=url)
    if filters is None or filters == "":
        # return datasets that have no filters
        return results.filter(filters=None)

    filter_count = len(filters)

    # first pass remove any results that do not have the same number of filters
    results_to_exclude = []
    for result in results:
        if len(result.filters.all()) != filter_count:
            results_to_exclude.append(result.id)

    for exclude in results_to_exclude:
        results = results.exclude(id=exclude)

    # now need to make sure filters match
    filter_dict = {}
    for filter_ in filters:
        name, value = filter_.split("=")
        filter_dict[name] = value

    results_to_exclude = []
    for result in results:
        for dataset_filter in result.filters.all():
            if filter_dict.get(dataset_filter.name) != dataset_filter.value:
                results_to_exclude.append(result.id)
                # no need to check the rest of the filters
                break

    for exclude in results_to_exclude:
        results = results.exclude(id=exclude)

    return results


def _get_values(value, prefix=""):
    values = value.replace(",", "").split("\n")

    prefixed_values = []
    for value in values:
        cleaned_value = value.lstrip(" ").strip(" ")
        if cleaned_value == "-":
            continue
        prefixed_values.append(prefix + cleaned_value)

    return prefixed_values
